import openpyxl as xl
from openpyxl.chart import BarChart,Reference

def process_workbook(filename):

    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    #for making a corrected cell in a new file that contain a new column that has the corrected values
    for row in range(2,sheet.max_row+1):
        cell = sheet.cell(row,3)
        corrected_prices = cell.value*0.9
        corrected_prices_cell = sheet.cell(row,4)
        corrected_prices_cell.value = corrected_prices

    sheet.cell(1,4).value = "correct price"


    #for creating a bar chart in our new excel sheet that contains the corrected price cell
    values = Reference(sheet,min_row = 2, max_row = sheet.max_row, min_col = 4, max_col = 4)
    chart = BarChart()
    chart.title = "Correct Prices Bar Chart"
    chart.add_data(values)
    sheet.add_chart(chart,'e2')

    #creating bar chart of our initial prices
    values1 = Reference(sheet,min_row = 2, max_row = sheet.max_row, min_col = 3, max_col = 3)
    chart1 = BarChart()
    chart1.title = "InCorrect Prices Bar Chart"
    chart1.add_data(values1)
    sheet.add_chart(chart1,'e20')


    #adding a chart for a row 
    values2 = Reference(sheet,min_col = 1, max_col = sheet.max_column, min_row=4,max_row=4)
    chart2 = BarChart()
    chart2.title = "Row 4 Values Bar Chart"
    chart2.add_data(values2)
    sheet.add_chart(chart2,'e38')
    wb.save(filename)